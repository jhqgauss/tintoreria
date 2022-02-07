from openpyxl import Workbook, load_workbook

libro = Workbook()

page = libro.active

companies = [['n1','n2','n3','n4','n5','n6','n7']]
encabezado = ['Item','Descripcion','Cant mano UMP','Nivel reorden','Codigo','Provedor','Pendientes','Programacion']

page.append(encabezado)

suma = 10+50

companies[0].append(suma)

for companie in companies:
    page.append(companie)
libro.save(filename="combinacion.xlsx")


# wb = load_workbook("archives/Niveles_inventario Convertido2.xlsx",read_only=True)
# wb.sheetnames
# ws = wb.active
#Extracts the subscriber number for accounts that went live last week.
# for row in ws.iter_rows(min_row=2,values_only=True):
#     for cell in row:
#         if cell == 21150715:
#             print(row)


# def buscarregistro():
#     for row in ws.iter_rows(min_row=2,values_only=True):
#         for cell in row:
#             if cell == 21150715:
#                 return (row)

# array1 = buscarregistro()

# print(array1)
# print(array1[2])