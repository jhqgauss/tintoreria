from ast import Not
# from crypt import methods
from dataclasses import dataclass
from turtle import pd
from xmlrpc.client import boolean
from flask import Flask,request,jsonify,json,render_template
from openpyxl import Workbook,load_workbook
import numpy as np
import os

from pandas import array

app = Flask(__name__)

libro = Workbook()

def buscarregistro(register,archive):
    for row in archive.iter_rows(min_row=2,values_only=True):
        for cell in row:
            if cell == register:
                return (row)


def archivocombinacion(data):
    page = libro.active
    encabezado = ['Item','Descripcion','Cant mano UMP','Nivel reorden','Codigo','Provedor','Pendientes','Programacion']
    page.append(encabezado)
    print("data de la funcion:",data)

    page.append(data)

    return 'insertado'




@app.route('/')
def index():
    return render_template('index.html')


@app.route('/file', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':

        if not os.path.exists('archives'):
            os.mkdir('archives')

        if len(os.listdir('archives')) >= 2:
            ruta="combination"
            tipo='ya se poseen los archivos necesarios. Visualizar combinacion'
        else:
            file  = request.files['file']
            file.save(os.path.join('archives',file.filename))
            ruta="/"
            tipo='Archivo añadido correctamente'

        return render_template('get_update.html',tipo=tipo,ruta=ruta)
    else:

        cobinacion = []
        ab = load_workbook(filename = 'combinacion.xlsx', read_only=True)
        cv = ab['Sheet']
        for r in cv.iter_rows(min_col=1, max_col=8, min_row=2, values_only=True):
                cobinacion.append(r)

    return render_template('index.html',data=cobinacion)

# la peticion put se reemplaza por get debido que html solo soporta post y get
@app.route('/file/<int:rowindex>', methods=['GET'])
def update(rowindex):
    ab = load_workbook(filename = 'combinacion.xlsx', read_only=False)
    cv = ab['Sheet']

    valor = request.args.get('programacion')
    celda = str(f'H{rowindex+1}')
    print(celda)

    cv[celda] = valor

    ab.save("combinacion.xlsx")

    ruta="/file"

    return render_template('get_update.html',tipo='Actualizado correctamente',ruta=ruta)

@app.route('/combination', methods=['GET'])
def combination():

    array1 = []
    array2 = []

    archivos = os.listdir('archives')
    for archivo in archivos:
        wb = load_workbook(filename = f'archives/{archivo}', read_only=True)
        ws = wb['in']
        for cell in ws.iter_rows(min_col=4, max_col=4, min_row=2, values_only=True):
            if cell[0] is not None and  type(cell[0]) == int:
                if archivo == 'Niveles_inventario Convertido.xlsx':
                    array1.append(cell[0])
                else:
                    array2.append(cell[0])
    print("array1",array1)
    print("array2",array2)

    coincidencias = np.intersect1d(array1,array2)
    print("coincidencias",coincidencias)

    # creacion archivo y insersion

    if os.path.exists('combinacion.xlsx'):
        os.remove('combinacion.xlsx')

    page = libro.active
            
    encabezado = ['Item','Descripcion','Cant mano UMP','Nivel reorden','Codigo','Provedor','Pendientes','Programacion']
    page.append(encabezado)

    for conc in coincidencias:
        data_fila = []
        for archivo in archivos:
            wb = load_workbook(filename = f'archives/{archivo}', read_only=True)
            ws = wb['in']
            reg = buscarregistro(conc,ws)
            if(archivo == 'Niveles_inventario Convertido.xlsx'):
                item = reg[3]
                descripcion = reg[4]
                cant_mano_UMP = reg[8]
                if cant_mano_UMP is None:
                    cant_mano_UMP = 0
                nivel_orden = reg[17]
                if nivel_orden is None:
                    nivel_orden = 0
            else:
                codigo = reg[3]
                provedor = reg[2]
                pendientes = reg[14]
                if pendientes is None:
                    pendientes = 0

        data_fila.append(item)
        data_fila.append(descripcion)
        data_fila.append(cant_mano_UMP)
        data_fila.append(nivel_orden)
        data_fila.append(codigo)
        data_fila.append(provedor)
        data_fila.append(pendientes)
        programacion = ((float(nivel_orden))-(float(cant_mano_UMP))-(float(pendientes)))
        print("nivel_orden",(float(nivel_orden)))
        print("cant_mano_UMP",(float(cant_mano_UMP)))
        print("pendientes",(float(pendientes)))
        data_fila.append(programacion)
        print("data_fila:",data_fila)
        page.append(data_fila)

    libro.save(filename="combinacion.xlsx")

    ruta="/file"

    return render_template('get_update.html',tipo='Combinacion correcta',ruta=ruta)



if __name__=='__main__':
    app.run(debug=True)